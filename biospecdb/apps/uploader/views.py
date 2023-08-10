from django.shortcuts import render
from .forms import FileUploadForm, DataInputForm
from openpyxl import load_workbook

def home(request):
    context = {'name': 'World'}
    render(request, 'Home.html', context)
    return upload_file(request)
    #return render(request, 'Home.html', context)


def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            #return render(request, 'UploadSuccess.html')
            return display_xlsx(request)
    else:
        form = FileUploadForm()
    return render(request, 'MetadataFileUpload.html', {'form': form})


def display_xlsx(request):
    workbook = load_workbook('./biospecdb/apps/uploader/uploads/METADATA_barauna2021ultrarapid.xlsx')
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    return render(request, 'MetadataDisplay.html', {'data': data})

def data_input_view(request):
    if request.method == 'POST':
        form = DataInputForm(request.POST, request.FILES)
        if form.is_valid():
            return render(request, 'DataInputForm_Success.html', {'form': form})
    else:
        form = DataInputForm()

    return render(request, 'DataInputForm.html', {'form': form})
